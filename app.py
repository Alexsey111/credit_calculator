from flask import Flask, render_template, request, jsonify, Response, send_file
import io
from datetime import datetime

app = Flask(__name__)

def build_amortization_schedule(loan_amount, years, interest_rate, prepayments=None, strategy='reduce_term'):
    prepayments = prepayments or []
    prepay_map = {int(p.get('month', 0)): float(p.get('amount', 0)) for p in prepayments if float(p.get('amount', 0)) > 0}

    monthly_rate = interest_rate / 100 / 12
    total_payments = years * 12

    if total_payments <= 0:
        raise ValueError('Срок должен быть положительным')

    # Рассчитываем первоначальный платеж
    if monthly_rate == 0:
        monthly_payment = loan_amount / total_payments
    else:
        factor = (1 + monthly_rate) ** total_payments
        monthly_payment = loan_amount * (monthly_rate * factor) / (factor - 1)

    schedule = []
    remaining_balance = loan_amount
    current_payment_index = 0

    while remaining_balance > 0 and current_payment_index < 1000 * total_payments:
        current_payment_index += 1
        interest_payment = remaining_balance * monthly_rate if monthly_rate > 0 else 0.0
        principal_payment = monthly_payment - interest_payment
        if principal_payment <= 0 and monthly_rate > 0:
            # защита от бесконечного цикла при экстремальных значениях
            principal_payment = 0.01
        remaining_balance = max(0.0, remaining_balance - principal_payment)

        # Досрочный платеж в этом месяце (после основного платежа)
        extra = prepay_map.get(current_payment_index, 0.0)
        if extra > 0 and remaining_balance > 0:
            remaining_balance = max(0.0, remaining_balance - extra)
            if strategy == 'reduce_payment' and remaining_balance > 0:
                # пересчитываем ежемесячный платеж на оставшийся срок
                payments_left = max(1, total_payments - current_payment_index)
                if monthly_rate == 0:
                    monthly_payment = remaining_balance / payments_left
                else:
                    factor_left = (1 + monthly_rate) ** payments_left
                    monthly_payment = remaining_balance * (monthly_rate * factor_left) / (factor_left - 1)
            # strategy == 'reduce_term' — платеж оставляем прежним, срок сократится автоматически

        schedule.append({
            'month': current_payment_index,
            'payment': round(principal_payment + interest_payment, 2),
            'principal': round(principal_payment, 2),
            'interest': round(interest_payment, 2),
            'extra': round(extra, 2),
            'remaining_balance': round(remaining_balance, 2)
        })

        if remaining_balance <= 0.005:
            remaining_balance = 0.0
            break

    total_paid = sum(r['payment'] for r in schedule) + sum(r['extra'] for r in schedule)
    overpayment = total_paid - loan_amount

    return schedule, monthly_payment, total_paid, overpayment

def calculate_mortgage(loan_amount, years, interest_rate, prepayments=None, strategy='reduce_term'):
    full_schedule, monthly_payment, total_payment, overpayment = build_amortization_schedule(
        loan_amount=loan_amount,
        years=years,
        interest_rate=interest_rate,
        prepayments=prepayments,
        strategy=strategy
    )
    short_schedule = full_schedule[:12]
    return {
        'monthly_payment': round(monthly_payment, 2),
        'total_payment': round(total_payment, 2),
        'overpayment': round(overpayment, 2),
        'overpayment_percentage': round((overpayment / loan_amount) * 100, 2) if loan_amount > 0 else 0.0,
        'payment_schedule': short_schedule,
        'full_schedule': full_schedule,
        'total_payments': len(full_schedule)
    }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/calculate', methods=['POST'])
def calculate():
    try:
        data = request.get_json(force=True)
        loan_amount = float(data['loan_amount'])
        years = int(data['years'])
        installment = bool(data.get('installment', False))
        interest_rate = 0.0 if installment else float(data.get('interest_rate', 0.0))
        prepayments = data.get('prepayments', [])
        strategy = data.get('strategy', 'reduce_term')

        if loan_amount <= 0 or years <= 0 or interest_rate < 0:
            return jsonify({'error': 'Все значения должны быть положительными'}), 400

        result = calculate_mortgage(loan_amount, years, interest_rate, prepayments, strategy)
        return jsonify(result)

    except (ValueError, KeyError, TypeError):
        return jsonify({'error': 'Неверные данные'}), 400
    except Exception:
        return jsonify({'error': 'Ошибка расчета'}), 500

@app.route('/download-csv')
def download_csv():
    try:
        loan_amount = float(request.args.get('loan_amount', ''))
        years = int(request.args.get('years', ''))
        installment = request.args.get('installment', 'false').lower() == 'true'
        interest_rate = 0.0 if installment else float(request.args.get('interest_rate', '0'))
        strategy = request.args.get('strategy', 'reduce_term')
        # простая форма: один досрочный платеж
        prepay_amount = float(request.args.get('prepay_amount', '0') or 0)
        prepay_month = int(request.args.get('prepay_month', '0') or 0)
        prepayments = []
        if prepay_amount > 0 and prepay_month > 0:
            prepayments.append({'month': prepay_month, 'amount': prepay_amount})

        result = calculate_mortgage(loan_amount, years, interest_rate, prepayments, strategy)
        schedule = result['full_schedule']

        lines = ['Месяц,Платеж,Основной долг,Проценты,Досрочный платеж,Остаток долга']
        for row in schedule:
            lines.append(f"{row['month']},{row['payment']},{row['principal']},{row['interest']},{row['extra']},{row['remaining_balance']}")
        csv_content = '\n'.join(lines)

        filename = f"mortgage_{int(loan_amount)}_{years}y_{interest_rate}pct_{strategy}.csv"
        return Response(csv_content, mimetype='text/csv; charset=utf-8', headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Exception:
        return jsonify({'error': 'Не удалось сформировать CSV'}), 500

@app.route('/download-xlsx')
def download_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        loan_amount = float(request.args.get('loan_amount', ''))
        years = int(request.args.get('years', ''))
        installment = request.args.get('installment', 'false').lower() == 'true'
        interest_rate = 0.0 if installment else float(request.args.get('interest_rate', '0'))
        strategy = request.args.get('strategy', 'reduce_term')
        prepay_amount = float(request.args.get('prepay_amount', '0') or 0)
        prepay_month = int(request.args.get('prepay_month', '0') or 0)
        prepayments = []
        if prepay_amount > 0 and prepay_month > 0:
            prepayments.append({'month': prepay_month, 'amount': prepay_amount})

        result = calculate_mortgage(loan_amount, years, interest_rate, prepayments, strategy)
        schedule = result['full_schedule']

        wb = Workbook()
        ws = wb.active
        ws.title = 'График платежей'

        headers = ['Месяц', 'Платеж', 'Основной долг', 'Проценты', 'Досрочный платеж', 'Остаток долга']
        ws.append(headers)
        header_fill = PatternFill('solid', fgColor='DDD6FE')
        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        for r in schedule:
            ws.append([r['month'], r['payment'], r['principal'], r['interest'], r['extra'], r['remaining_balance']])

        # Итоги на отдельной вкладке
        ws2 = wb.create_sheet('Итоги')
        ws2.append(['Ежемесячный платеж', result['monthly_payment']])
        ws2.append(['Всего выплат', result['total_payment']])
        ws2.append(['Переплата', result['overpayment']])
        ws2.append(['% переплаты', result['overpayment_percentage']])

        for sheet in (ws, ws2):
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[col_letter].width = max(12, min(32, max_length + 2))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"mortgage_{int(loan_amount)}_{years}y_{interest_rate}pct_{strategy}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception:
        return jsonify({'error': 'Не удалось сформировать Excel'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
